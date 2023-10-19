import openpyxl
from datetime import datetime

# Função para converter string de data para objeto datetime
def str_para_data(data_str):
    return datetime.strptime(data_str, "%d/%m/%Y")

# Função para converter string de competência para objeto datetime
def str_para_competencia(comp_str):
    return datetime.strptime(comp_str, "%m/%Y")

def planilha(dados):

    planilha = 'Z:\RPA\Simples Nacional\BD_Simples_Nacional\Acompanhamento tributário.xlsx'

    # Carregar a planilha existente
    workbook = openpyxl.load_workbook(planilha)

    # Selecionar a aba 'bd'
    sheet = workbook['bd']

    # Extrair CNPJ e competência dos dados
    
    cnpj = dados[0]
    competencia_str = dados[3]
    competencia = str_para_data(competencia_str)
    dados[3] = competencia
    
    # Flag para verificar se a linha foi atualizada
    linha_atualizada = False
    
    # Iterar sobre as linhas existentes na planilha
    for row_num in range(2, sheet.max_row + 1):
        
        # Verificar se o CNPJ e a competência correspondem
        if sheet.cell(row=row_num, column=1).value == cnpj and \
        sheet.cell(row=row_num, column=4).value == competencia:
            
            # Atualizar a linha
            for col_num, valor in enumerate(dados, start=1):
                sheet.cell(row=row_num, column=col_num, value=valor)
                # Copiar a formatação da célula acima
                sheet.cell(row=row_num, column=col_num)._style = sheet.cell(row=row_num-1, column=col_num)._style
            linha_atualizada = True
            break
    
    # Se a linha não foi atualizada, adicionar uma nova linha
    if not linha_atualizada:
        proxima_linha = sheet.max_row + 1
        for col_num, valor in enumerate(dados, start=1):
            sheet.cell(row=proxima_linha, column=col_num, value=valor)
            # Copiar a formatação da célula acima
            sheet.cell(row=proxima_linha, column=col_num)._style = sheet.cell(row=proxima_linha-1, column=col_num)._style

    # Salvar a planilha
    workbook.save(planilha)

    return True

def atualizar_valor_fopag(cnpj, competencia, novo_valor):
    planilha = 'Z:\RPA\Simples Nacional\BD_Simples_Nacional\Acompanhamento tributário.xlsx'

    # Carregar a planilha existente
    workbook = openpyxl.load_workbook(planilha)

    # Selecionar a aba 'bd'
    sheet = workbook['bd']

    # Flag para verificar se a linha foi atualizada
    linha_atualizada = False

    # Iterar sobre as linhas existentes na planilha
    for row_num in range(2, sheet.max_row + 1):
        # Verificar se o CNPJ e a competência correspondem
        if sheet.cell(row=row_num, column=1).value == cnpj and \
        sheet.cell(row=row_num, column=4).value == competencia:
            # Atualizar a coluna "valor_fopag"
            sheet.cell(row=row_num, column=10).value = novo_valor
            sheet.cell(row=row_num, column=10)._style = sheet.cell(row=row_num-1, column=10)._style
            linha_atualizada = True
            break

    # Se a linha não foi atualizada, você pode optar por adicionar uma nova linha ou retornar um erro
    if not linha_atualizada:
        print(f"Não foi possível encontrar uma linha com o CNPJ {cnpj} e a competência {competencia}.")
        return False

    # Salvar a planilha
    workbook.save(planilha)

    return True